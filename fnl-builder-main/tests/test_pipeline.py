from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]

from fnl_builder.config import InputPaths, PipelineConfig, RunState
from fnl_builder.pipeline import integrate_stage, render_stage
from fnl_builder.shared.types import (
    GuestRecord,
    InquiryKey,
    IntegrationResult,
    MessageListData,
    ParseResult,
    PassengerData,
    RewriteStats,
    RoomingData,
    TourHeaderData,
)


def _build_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "HTL rooming"
    ws.cell(row=1, column=1).value = "Tour REF"
    ws.cell(row=2, column=1).value = "Tour NAME"
    ws.cell(row=3, column=1).value = "Total PAX"
    ws.cell(row=4, column=1).value = "Total Rooms"
    ws.cell(row=10, column=1).value = "room type"
    ws.cell(row=10, column=2).value = "number"
    ws.cell(row=10, column=3).value = "inquiryno"
    ws.cell(row=10, column=4).value = "familyname"
    ws.cell(row=10, column=5).value = "givenname"
    ws.cell(row=10, column=6).value = "remarks"
    wb.save(path)


def _make_state(tmp_path: Path) -> RunState:
    template_path = tmp_path / "template.xlsx"
    _build_template(template_path)
    config = PipelineConfig(
        llm_provider="none",
        input_paths=InputPaths(
            rooming=tmp_path / "dummy_rl.pdf",
            template=template_path,
            output=tmp_path / "out.xlsx",
            audit=tmp_path / "audit.json",
        ),
    )
    return RunState.from_config(config)


def test_integrate_stage_empty_input() -> None:
    parsed = ParseResult(
        rooming=RoomingData.empty(),
        passenger=PassengerData.empty(),
        messagelist=MessageListData.empty(),
        tour_header=TourHeaderData.empty(),
    )
    config = PipelineConfig(llm_provider="none")
    state = RunState.from_config(config)
    result = integrate_stage(parsed, state)
    assert result.guests == []
    assert result.stats.candidates == 0


def test_integrate_stage_with_guests() -> None:
    guest = GuestRecord(
        inquiry=InquiryKey(main="10000001"),
        full_name="HANKYU TARO",
        family_name="HANKYU",
        given_name="TARO",
        room_type="TWN",
        room_group_id="g1",
    )
    rooming = RoomingData(guests=[guest])
    parsed = ParseResult(
        rooming=rooming,
        passenger=PassengerData.empty(),
        messagelist=MessageListData.empty(),
        tour_header=TourHeaderData.empty(),
    )
    config = PipelineConfig(llm_provider="none")
    state = RunState.from_config(config)
    result = integrate_stage(parsed, state)
    assert len(result.guests) == 1
    assert result.guests[0].family_name == "HANKYU"


def test_render_stage_writes_excel(tmp_path: Path) -> None:
    state = _make_state(tmp_path)
    guest = GuestRecord(
        inquiry=InquiryKey(main="10000001"),
        full_name="HANKYU TARO",
        family_name="HANKYU",
        given_name="TARO",
        room_type="TWN",
        room_number="1",
        room_group_id="g1",
    )
    rooming = RoomingData(tour_ref="E417", tour_name="TEST TOUR", declared_total_pax=1)
    integrated = IntegrationResult(
        guests=[guest],
        companion_groups={},
        stats=RewriteStats(),
    )
    result = render_stage(integrated, state, rooming=rooming)
    assert result.output_path.exists()
    assert isinstance(result.audit, type(state.audit))


def test_render_stage_writes_audit_json(tmp_path: Path) -> None:
    state = _make_state(tmp_path)
    integrated = IntegrationResult(
        guests=[],
        companion_groups={},
        stats=RewriteStats(),
    )
    rooming = RoomingData.empty()
    render_stage(integrated, state, rooming=rooming)
    audit_path = tmp_path / "audit.json"
    assert audit_path.exists()
