from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

import pytest

from fnl_builder.render.excel import (
    _inquiry_display_value,
    _normalize_header,
    find_guest_header_row,
    render_final_list_workbook,
    write_guest_rows,
)
from fnl_builder.render.excel_text import (
    display_width,
    sanitize_excel_text,
)
from fnl_builder.render.remarks_format import format_guest_remarks
from fnl_builder.shared.errors import InputError
from fnl_builder.shared.types import GuestRecord, InquiryKey


def _build_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "HTL rooming"
    worksheet.cell(row=1, column=1).value = "Tour REF"
    worksheet.cell(row=2, column=1).value = "Tour NAME"
    worksheet.cell(row=3, column=1).value = "Total PAX"
    worksheet.cell(row=4, column=1).value = "Total Rooms"
    worksheet.cell(row=10, column=1).value = "room type"
    worksheet.cell(row=10, column=2).value = "number"
    worksheet.cell(row=10, column=3).value = "inquiryno"
    worksheet.cell(row=10, column=4).value = "familyname"
    worksheet.cell(row=10, column=5).value = "givenname"
    worksheet.cell(row=10, column=6).value = "remarks"
    workbook.save(path)


def test_excel_render_writes_summary_cells(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    out_path = tmp_path / "out.xlsx"
    _build_template(template_path)

    render_final_list_workbook(
        template_path=template_path,
        out_path=out_path,
        tour_ref="E417",
        tour_name="SAMPLE TOUR",
        total_pax=2,
        rooms_by_type={"TWN": 1},
        guests=[],
        companion_groups={},
        issues=[],
    )

    loaded = load_workbook(out_path)
    ws = loaded["HTL rooming"]
    assert ws.cell(row=1, column=2).value == "E417"
    assert ws.cell(row=2, column=2).value == "SAMPLE TOUR"
    assert ws.cell(row=3, column=2).value == "2PAX"
    assert ws.cell(row=4, column=2).value == "1TWN"


def test_excel_render_formats_remarks_with_category_lines_and_wrap(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    out_path = tmp_path / "out.xlsx"
    _build_template(template_path)

    guest = GuestRecord(
        inquiry=InquiryKey(main="10000001"),
        full_name="HANKYU TARO",
        family_name="HANKYU",
        given_name="TARO",
        room_type="TWN",
        room_number="1",
        remarks_parts=[
            "[同室] HANKYU HANAKO様とTWN同室",
            "[同行GRP別室] HANSHIN JIRO様(TSU/No.1)",
            "[medical] 糖尿病対応が必要",
            "[medical] 注射器は機内持ち込み",
            "[baggage] 血糖測定器具の持ち込みあり",
            "[group] 同行グループ（別問合せ番号）として同一扱い",
            "フルーツバスケット手配; ベジ対応希望",
        ],
    )

    render_final_list_workbook(
        template_path=template_path,
        out_path=out_path,
        tour_ref="E417",
        tour_name="SAMPLE TOUR",
        total_pax=1,
        rooms_by_type={"TWN": 1},
        guests=[guest],
        companion_groups={},
        issues=[],
    )

    loaded = load_workbook(out_path)
    ws = loaded["HTL rooming"]
    remark_cell = ws.cell(row=11, column=6)
    remark_text = remark_cell.value
    assert isinstance(remark_text, str)
    assert remark_text == format_guest_remarks(guest.remarks_parts)
    assert "[medical] 糖尿病対応が必要\n  注射器は機内持ち込み" in remark_text
    assert "[other] フルーツバスケット手配\n  ベジ対応希望" in remark_text
    assert "\n" in remark_text
    assert ";" not in remark_text
    assert "；" not in remark_text
    assert remark_cell.alignment.wrap_text is True
    assert (ws.column_dimensions["F"].width or 0) >= 24
    assert (ws.row_dimensions[11].height or 0) > 15


def test_find_guest_header_row_basic() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.cell(row=8, column=1).value = "room type"
    ws.cell(row=8, column=2).value = "number"
    ws.cell(row=8, column=3).value = "inquiryno"
    ws.cell(row=8, column=4).value = "familyname"
    ws.cell(row=8, column=5).value = "givenname"
    ws.cell(row=8, column=6).value = "remarks"

    found = find_guest_header_row(ws)
    assert found is not None
    row, cols = found
    assert row == 8
    assert cols["inquiry"] == 3
    assert cols["remarks"] == 6


def test_find_guest_header_row_missing() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.cell(row=1, column=1).value = "not headers"
    assert find_guest_header_row(ws) is None


def test_sanitize_excel_text_formula_prefix() -> None:
    assert sanitize_excel_text("=SUM") == "'=SUM"


def test_inquiry_display_value_numeric() -> None:
    guest = GuestRecord(
        inquiry=InquiryKey(main="00012345"),
        full_name="A B",
        family_name="A",
        given_name="B",
    )
    assert _inquiry_display_value(guest) == 12345


def test_inquiry_display_value_with_branch() -> None:
    guest = GuestRecord(
        inquiry=InquiryKey(main="00012345", branch="2"),
        full_name="A B",
        family_name="A",
        given_name="B",
    )
    assert _inquiry_display_value(guest) == "12345-2"


def test_sanitize_excel_text_variants() -> None:
    assert sanitize_excel_text("=CMD()") == "'=CMD()"
    assert sanitize_excel_text("+1") == "'+1"
    assert sanitize_excel_text("-1") == "'-1"
    assert sanitize_excel_text("@SUM") == "'@SUM"
    assert sanitize_excel_text("\t=foo") == "'\t=foo"
    assert sanitize_excel_text("normal text") == "normal text"
    assert sanitize_excel_text("") == ""


def test_display_width_ascii_cjk_and_mixed() -> None:
    assert display_width("abc") == 3
    assert display_width("漢字") == 4
    assert display_width("ABカナ") == 6


def test_normalize_header_variants() -> None:
    assert _normalize_header("Room Type") == "roomtype"
    assert _normalize_header("room_number") == "roomnumber"
    assert _normalize_header("  INQUIRY no  ") == "inquiryno"


def test_find_guest_header_row_inquiry_alias_toi_awase_no() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.cell(row=6, column=1).value = "room type"
    ws.cell(row=6, column=2).value = "number"
    ws.cell(row=6, column=3).value = "問い合わせNo"
    ws.cell(row=6, column=4).value = "familyname"
    ws.cell(row=6, column=5).value = "givenname"
    ws.cell(row=6, column=6).value = "remarks"

    found = find_guest_header_row(ws)
    assert found is not None
    row, cols = found
    assert row == 6
    assert cols["inquiry"] == 3


def test_find_guest_header_row_inquiry_alias_toiawase_no() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.cell(row=7, column=1).value = "room type"
    ws.cell(row=7, column=2).value = "number"
    ws.cell(row=7, column=3).value = "問合せNO"
    ws.cell(row=7, column=4).value = "familyname"
    ws.cell(row=7, column=5).value = "givenname"
    ws.cell(row=7, column=6).value = "remarks"

    found = find_guest_header_row(ws)
    assert found is not None
    row, cols = found
    assert row == 7
    assert cols["inquiry"] == 3


def test_write_guest_rows_missing_required_column_raises() -> None:
    workbook = Workbook()
    ws = workbook.active
    cols = {"inquiry": 1, "family_name": 2}  # missing given_name, remarks
    with pytest.raises(InputError, match="missing required columns"):
        write_guest_rows(ws, 1, cols, [], tour_ref=None)
