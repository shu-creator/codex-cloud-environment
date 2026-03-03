"""CI-safe E2E smoke test with synthetic fixtures."""
from __future__ import annotations

import json
from pathlib import Path

from fnl_builder.config import InputPaths, PipelineConfig
from fnl_builder.parse.input_extract import PdfExtractionMeta
from fnl_builder.pipeline import run

_FIXTURE_DIR = Path(__file__).parent / "fixtures" / "smoke_minimum"
_TEMPLATE_PATH = Path(__file__).parent / "fixtures" / "e417_1008" / "template.xlsx"


def _load_expected() -> dict[str, object]:
    return json.loads((_FIXTURE_DIR / "expected.json").read_text(encoding="utf-8"))  # type: ignore[return-value]


def _load_rooming_text() -> str:
    return (_FIXTURE_DIR / "rooming.txt").read_text(encoding="utf-8")


def _load_passenger_text() -> str:
    return (_FIXTURE_DIR / "passenger.txt").read_text(encoding="utf-8")


def _mock_extract_pdf_text(path: Path) -> tuple[str, PdfExtractionMeta]:
    if path.name == "rooming.pdf":
        return _load_rooming_text(), PdfExtractionMeta(method="mock", total_pages=1)
    if path.name == "passenger.pdf":
        return _load_passenger_text(), PdfExtractionMeta(method="mock", total_pages=1)
    raise AssertionError(f"Unexpected PDF path: {path}")


def test_pipeline_smoke_minimum(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr("fnl_builder.pipeline.extract_pdf_text", _mock_extract_pdf_text)

    rooming_path = tmp_path / "rooming.pdf"
    passenger_path = tmp_path / "passenger.pdf"
    rooming_path.write_bytes(b"")
    passenger_path.write_bytes(b"")

    config = PipelineConfig(
        llm_provider="none",
        input_paths=InputPaths(
            rooming=rooming_path,
            passenger=passenger_path,
            messagelist=_FIXTURE_DIR / "messagelist.csv",
            template=_TEMPLATE_PATH,
            output=tmp_path / "out.xlsx",
            audit=tmp_path / "audit.json",
        ),
    )

    result = run(config)
    expected = _load_expected()

    assert len(result.audit.issues) == len(expected["issue_codes"])
    assert sorted({issue.code for issue in result.audit.issues}) == expected["issue_codes"]

    from openpyxl import load_workbook  # type: ignore[import-untyped]
    from fnl_builder.render.excel import find_guest_header_row

    workbook = load_workbook(result.output_path, data_only=True)
    worksheet = workbook.active

    header = find_guest_header_row(worksheet)
    assert header is not None
    header_row, cols = header

    family_names: list[str] = []
    room_types: list[str] = []
    course_codes: list[str] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        room_type = str(worksheet.cell(row=row, column=cols["room_type"]).value or "").strip()
        course_code = ""
        if "tour_ref" in cols:
            course_code = str(worksheet.cell(row=row, column=cols["tour_ref"]).value or "").strip()
        family_name = str(worksheet.cell(row=row, column=cols["family_name"]).value or "").strip()
        given_name = str(worksheet.cell(row=row, column=cols["given_name"]).value or "").strip()
        if not room_type and not family_name and not given_name:
            continue
        if family_name:
            family_names.append(family_name.replace("MR. ", "").replace("MS. ", "").strip())
        if room_type:
            room_types.append(room_type)
        if course_code:
            course_codes.append(course_code)

    assert len(family_names) == expected["guest_count"]
    assert family_names == expected["family_names"]
    assert room_types == expected["room_types"]
    assert course_codes == expected["course_codes"]
    assert result.output_path.exists()
