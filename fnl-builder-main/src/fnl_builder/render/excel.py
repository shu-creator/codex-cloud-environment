from __future__ import annotations

import re
from copy import copy
from importlib.abc import Traversable
from importlib.resources import files
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Border, Side  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from fnl_builder.render.excel_text import (
    apply_wrap_alignment,
    autosize_wrapped_text_area,
    sanitize_excel_text,
)
from fnl_builder.render.remarks_format import format_guest_remarks
from fnl_builder.shared.errors import InputError
from fnl_builder.shared.io import atomic_save_workbook
from fnl_builder.shared.text import collapse_ws, normalize_inquiry_main
from fnl_builder.shared.types import GuestRecord, Issue

def default_template_ref() -> Traversable:
    """Return a Traversable reference to the bundled default template.

    Use with ``importlib.resources.as_file`` at the call site::

        with as_file(default_template_ref()) as path:
            ...
    """
    return files("fnl_builder.render").joinpath("template.xlsx")


_REQUIRED_HEADERS = {"room_type", "room_number", "inquiry", "family_name", "given_name", "remarks"}
_HEADER_ALIASES: dict[str, set[str]] = {
    "room_type": {"roomtype", "room type"},
    "room_number": {"number", "roomno", "roomnumber"},
    "tour_ref": {"コース番号", "courseno", "tourref", "tourref."},
    "inquiry": {"問い合わせno", "問合せno", "問合せno.", "inquiryno", "inquiryno"},
    "family_name": {"familyname"},
    "given_name": {"givenname"},
    "remarks": {"remarks", "備考"},
    "passport_no": {"pptno", "passportno", "passport"},
    "issue_date": {"issuedate", "issue"},
    "expiry_date": {"expirydate", "expiredate", "expiry"},
}


def _set_labeled_value(ws: Worksheet, label: str, value: object) -> bool:
    target = _find_label_cell(ws, label)
    if not target:
        return False
    row, col = target
    ws.cell(row=row, column=col + 1).value = value
    return True


def set_labeled_value(ws: Worksheet, label: str, value: object) -> bool:
    return _set_labeled_value(ws, label, value)


def _find_label_cell(ws: Worksheet, label: str) -> tuple[int, int] | None:
    want = collapse_ws(label).lower()
    for row in range(1, 60):
        for col in range(1, 25):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str):
                got = collapse_ws(value).lower()
                if got == want or want in got:
                    return row, col
    return None


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src, dst = ws.cell(row=src_row, column=col), ws.cell(row=dst_row, column=col)
        dst.number_format = src.number_format
        dst.protection, dst.alignment = copy(src.protection), copy(src.alignment)
        dst.fill, dst.font, dst.border = copy(src.fill), copy(src.font), copy(src.border)


def _unmerge_cells_in_range(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> list[str]:
    to_unmerge: list[tuple[str, int, int, int, int]] = []
    warnings: list[str] = []
    for merged_range in ws.merged_cells.ranges:
        overlaps = (
            merged_range.min_row <= max_row
            and merged_range.max_row >= min_row
            and merged_range.min_col <= max_col
            and merged_range.max_col >= min_col
        )
        if not overlaps:
            continue
        fully_contained = (
            merged_range.min_row >= min_row
            and merged_range.max_row <= max_row
            and merged_range.min_col >= min_col
            and merged_range.max_col <= max_col
        )
        if fully_contained:
            to_unmerge.append(
                (merged_range.coord, merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
            )
        else:
            warnings.append(
                f"Merged cell range {merged_range.coord} partially overlaps guest data area; not unmerged to preserve template layout."
            )
    for coord, mr_min_row, mr_max_row, mr_min_col, mr_max_col in to_unmerge:
        top_left = ws.cell(row=mr_min_row, column=mr_min_col)
        saved = (
            copy(top_left.font),
            copy(top_left.fill),
            copy(top_left.border),
            copy(top_left.alignment),
            copy(top_left.protection),
            top_left.number_format,
        )
        ws.unmerge_cells(coord)
        for row in range(mr_min_row, mr_max_row + 1):
            for col in range(mr_min_col, mr_max_col + 1):
                if row == mr_min_row and col == mr_min_col:
                    continue
                cell = ws.cell(row=row, column=col)
                cell.font, cell.fill, cell.border = copy(saved[0]), copy(saved[1]), copy(saved[2])
                cell.alignment, cell.protection, cell.number_format = copy(saved[3]), copy(saved[4]), saved[5]
    return warnings


def _build_room_group_mappings(
    guests: list[GuestRecord],
) -> tuple[dict[str, str], dict[str, tuple[str, str | None]], dict[str, set[str]]]:
    inq_to_room_group: dict[str, str] = {}
    room_group_info: dict[str, tuple[str, str | None]] = {}
    room_group_inquiries: dict[str, set[str]] = {}
    for guest in guests:
        if not guest.room_group_id:
            continue
        inquiry_key = normalize_inquiry_main(guest.inquiry.main)
        inq_to_room_group[inquiry_key] = guest.room_group_id
        if guest.room_group_id not in room_group_info and guest.room_type:
            room_group_info[guest.room_group_id] = (guest.room_type, guest.room_number)
        room_group_inquiries.setdefault(guest.room_group_id, set()).add(inquiry_key)
    return inq_to_room_group, room_group_info, room_group_inquiries


def _build_companion_rows(
    companion_groups: dict[str, set[str]] | None,
    inq_to_room_group: dict[str, str],
    room_group_info: dict[str, tuple[str, str | None]],
    room_group_inquiries: dict[str, set[str]],
) -> dict[str, str]:
    companion_rows: dict[str, str] = {}
    if not companion_groups:
        return companion_rows
    for room_group_id, inquiries in room_group_inquiries.items():
        for inquiry in inquiries:
            if inquiry not in companion_groups:
                continue
            for companion_inquiry in companion_groups[inquiry]:
                companion_room_group = inq_to_room_group.get(companion_inquiry)
                if not companion_room_group or companion_room_group == room_group_id:
                    continue
                info = room_group_info.get(companion_room_group)
                if not info:
                    continue
                room_type, room_number = info
                companion_rows[room_group_id] = (
                    f"{room_type}/No.{room_number}と同行GRP" if room_number else f"{room_type}/同行GRP"
                )
                break
            if room_group_id in companion_rows:
                break
    return companion_rows


def _ensure_row_capacity(ws: Worksheet, *, needed_last_row: int, template_style_row: int, max_col: int) -> None:
    if needed_last_row <= ws.max_row:
        return
    for row_number in range(ws.max_row + 1, needed_last_row + 1):
        ws.insert_rows(row_number)
        _copy_row_style(ws, template_style_row, row_number, max_col)


def _abort_on_merged_cell_overlap(
    ws: Worksheet, *, first_row: int, needed_last_row: int, max_col: int, issues: list[Issue] | None
) -> None:
    warnings = _unmerge_cells_in_range(ws, first_row, needed_last_row, 1, max_col)
    if not warnings:
        return
    if issues is not None:
        issues.extend(
            Issue(level="error", code="merged_cell_partial_overlap", message=warning) for warning in warnings
        )
    raise InputError(
        "Template contains merged cells that partially overlap the guest data area. "
        "Please remove or adjust these merged cells in the template: " + "; ".join(warnings)
    )


def _write_companion_row(ws: Worksheet, *, row: int, cols: dict[str, int], companion_text: str) -> None:
    for col_name, col_idx in cols.items():
        ws.cell(row=row, column=col_idx).value = sanitize_excel_text(companion_text) if col_name == "remarks" else None
        if col_name == "remarks":
            apply_wrap_alignment(ws, row=row, col=col_idx)


def _inquiry_display_value(guest: GuestRecord) -> str | int:
    inquiry_main = normalize_inquiry_main(guest.inquiry.main)
    if guest.inquiry.branch:
        return f"{inquiry_main}-{guest.inquiry.branch}"
    return int(inquiry_main) if inquiry_main.isdigit() else inquiry_main


def _title_prefix_from_full_name(full_name: str) -> str:
    match = re.match(r"^(MR\.?|MS\.?|MRS\.?|MISS\.?)\s+", full_name, re.IGNORECASE)
    return "" if not match else match.group(1).upper().rstrip(".") + ". "


def _safe_str(value: str | None) -> str:
    return sanitize_excel_text(value or "")


def _write_guest_row(
    ws: Worksheet, *, row: int, cols: dict[str, int], guest: GuestRecord, show_room: bool, tour_ref: str | None
) -> None:
    if "room_type" in cols:
        ws.cell(row=row, column=cols["room_type"]).value = guest.room_type if show_room else None
    if "room_number" in cols:
        room_number_value: str | int | None
        room_number_value = int(guest.room_number) if guest.room_number and guest.room_number.isdigit() else guest.room_number
        ws.cell(row=row, column=cols["room_number"]).value = room_number_value if show_room else None
    if "tour_ref" in cols:
        ws.cell(row=row, column=cols["tour_ref"]).value = _safe_str(guest.course_code or tour_ref or "")
    ws.cell(row=row, column=cols["inquiry"]).value = _inquiry_display_value(guest)
    ws.cell(row=row, column=cols["family_name"]).value = _safe_str(f"{_title_prefix_from_full_name(guest.full_name)}{guest.family_name}".strip())
    ws.cell(row=row, column=cols["given_name"]).value = _safe_str(guest.given_name)
    ws.cell(row=row, column=cols["remarks"]).value = sanitize_excel_text(format_guest_remarks(guest.remarks_parts))
    apply_wrap_alignment(ws, row=row, col=cols["remarks"])
    if "passport_no" in cols:
        ws.cell(row=row, column=cols["passport_no"]).value = _safe_str(guest.passport_no or "")
    if "issue_date" in cols:
        ws.cell(row=row, column=cols["issue_date"]).value = _safe_str(guest.issue_date or "")
    if "expiry_date" in cols:
        ws.cell(row=row, column=cols["expiry_date"]).value = _safe_str(guest.expiry_date or "")


def _apply_room_borders(
    ws: Worksheet, *, row: int, max_col: int, show_room: bool, room_top_border: Border, no_border: Border
) -> None:
    for col_idx in range(1, max_col + 1):
        ws.cell(row=row, column=col_idx).border = room_top_border if show_room else no_border


def _clear_rows_below(
    ws: Worksheet, *, start_row: int, end_row: int, cols: dict[str, int]
) -> None:
    no_border = Border()
    col_indices = sorted(cols.values())
    for clear_row in range(start_row, end_row + 1):
        for col_idx in col_indices:
            cell = ws.cell(row=clear_row, column=col_idx)
            if cell.value is not None:
                cell.value = None
            cell.border = no_border


def write_guest_rows(
    ws: Worksheet,
    header_row: int,
    cols: dict[str, int],
    guests: list[GuestRecord],
    *,
    tour_ref: str | None,
    companion_groups: dict[str, set[str]] | None = None,
    issues: list[Issue] | None = None,
) -> None:
    _WRITE_REQUIRED = {"inquiry", "family_name", "given_name", "remarks"}
    missing = _WRITE_REQUIRED - set(cols.keys())
    if missing:
        raise InputError(f"Template missing required columns: {missing}")
    first_row, max_col = header_row + 1, max(cols.values())
    template_last_row = ws.max_row if ws.max_row >= first_row else first_row
    inq_to_room_group, room_group_info, room_group_inquiries = _build_room_group_mappings(guests)
    companion_rows = _build_companion_rows(companion_groups, inq_to_room_group, room_group_info, room_group_inquiries)
    total_data_rows = len(guests) + len(companion_rows)
    needed_last_row = first_row + total_data_rows - 1 if total_data_rows > 0 else first_row
    _ensure_row_capacity(ws, needed_last_row=needed_last_row, template_style_row=first_row, max_col=max_col)
    _abort_on_merged_cell_overlap(ws, first_row=first_row, needed_last_row=needed_last_row, max_col=max_col, issues=issues)

    seen_room_group: set[str] = set()
    current_row, room_top_border, no_border = first_row, Border(top=Side(style="thin", color="999999")), Border()
    for idx, guest in enumerate(guests):
        show_room = guest.room_group_id not in seen_room_group
        if guest.room_group_id:
            seen_room_group.add(guest.room_group_id)
        _write_guest_row(ws, row=current_row, cols=cols, guest=guest, show_room=show_room, tour_ref=tour_ref)
        _apply_room_borders(
            ws, row=current_row, max_col=max_col, show_room=show_room, room_top_border=room_top_border, no_border=no_border
        )
        current_row += 1
        is_last = idx == len(guests) - 1 or guests[idx + 1].room_group_id != guest.room_group_id
        if is_last and guest.room_group_id and guest.room_group_id in companion_rows:
            _write_companion_row(ws, row=current_row, cols=cols, companion_text=companion_rows[guest.room_group_id])
            current_row += 1

    last_written_row = current_row - 1 if current_row > first_row else first_row
    clear_start = first_row if total_data_rows == 0 else last_written_row + 1
    clear_end = max(needed_last_row, template_last_row)
    if clear_start <= clear_end:
        _clear_rows_below(ws, start_row=clear_start, end_row=clear_end, cols=cols)
    autosize_wrapped_text_area(
        ws, col=cols["remarks"], first_row=first_row, last_row=last_written_row, template_style_row=first_row
    )


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s_]+", "", collapse_ws(value).lower())


def _header_field(normalized: str) -> str | None:
    for field, aliases in _HEADER_ALIASES.items():
        if normalized in aliases:
            return field
    return None


def find_guest_header_row(ws: Worksheet) -> tuple[int, dict[str, int]] | None:
    for row in range(1, 250):
        headers: dict[str, int] = {}
        for col in range(1, 80):
            cell_value = ws.cell(row=row, column=col).value
            if isinstance(cell_value, str):
                field = _header_field(_normalize_header(cell_value))
                if field:
                    headers[field] = col
        if _REQUIRED_HEADERS.issubset(headers):
            return row, headers
    return None


def render_final_list_workbook(
    *,
    template_path: Path,
    out_path: Path,
    tour_ref: str | None,
    tour_name: str | None,
    total_pax: int,
    rooms_by_type: dict[str, int],
    guests: list[GuestRecord],
    companion_groups: dict[str, set[str]] | None,
    issues: list[Issue] | None,
) -> None:
    workbook = load_workbook(template_path)
    if "HTL rooming" not in workbook.sheetnames:
        raise InputError("Template workbook must contain sheet 'HTL rooming'")
    worksheet = workbook["HTL rooming"]
    set_labeled_value(worksheet, "Tour REF", sanitize_excel_text(tour_ref or ""))
    set_labeled_value(worksheet, "Tour NAME", sanitize_excel_text(tour_name or ""))
    set_labeled_value(worksheet, "Total PAX", f"{total_pax}PAX")
    room_type_order = {"TWN": 0, "DBL": 1, "TPL": 2, "TRP": 2, "TSU": 3, "SGL": 4}
    out_rooms_text = "+".join(
        f"{count}{room_type}"
        for room_type, count in sorted(rooms_by_type.items(), key=lambda x: (room_type_order.get(x[0], 5), x[0]))
    )
    set_labeled_value(worksheet, "Total Rooms", out_rooms_text)
    header = find_guest_header_row(worksheet)
    if not header:
        raise InputError("Could not locate Guest list header row in template 'HTL rooming'")
    header_row, columns = header
    write_guest_rows(
        worksheet, header_row, columns, guests, tour_ref=tour_ref, companion_groups=companion_groups, issues=issues
    )
    atomic_save_workbook(out_path, workbook.save)
