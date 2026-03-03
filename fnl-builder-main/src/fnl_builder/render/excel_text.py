"""Text sanitization and display helpers for Excel rendering."""
from __future__ import annotations

import math
import unicodedata
from copy import copy

from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]


def sanitize_excel_text(text: str) -> str:
    """Prefix formula-like strings to prevent injection."""
    if text and text.lstrip()[:1] in {"=", "+", "-", "@"}:
        return "'" + text
    return text


def display_width(text: str) -> int:
    """Return visual width treating East-Asian wide chars as 2."""
    return sum(2 if unicodedata.east_asian_width(ch) in {"W", "F"} else 1 for ch in text)


def apply_wrap_alignment(ws: Worksheet, *, row: int, col: int) -> None:
    """Enable text wrapping on a cell."""
    cell = ws.cell(row=row, column=col)
    alignment = copy(cell.alignment)
    alignment.wrap_text, alignment.vertical = True, alignment.vertical or "top"
    cell.alignment = alignment


def autosize_wrapped_text_area(
    ws: Worksheet,
    *,
    col: int,
    first_row: int,
    last_row: int,
    template_style_row: int,
    min_width: int = 24,
    max_width: int = 120,
) -> None:
    """Auto-size column width and row heights for wrapped text."""
    col_letter, max_line_width = get_column_letter(col), 0
    values_by_row: dict[int, str] = {}
    for row in range(first_row, last_row + 1):
        value = ws.cell(row=row, column=col).value
        if not isinstance(value, str) or not value.strip():
            continue
        values_by_row[row] = value
        for line in value.splitlines() or [""]:
            max_line_width = max(max_line_width, display_width(line))
    if not values_by_row:
        return
    current_width = ws.column_dimensions[col_letter].width or 0
    target_width = min(max(max_line_width, min_width), max_width)
    if target_width > current_width:
        ws.column_dimensions[col_letter].width = float(target_width)
    base_height = ws.row_dimensions[template_style_row].height or ws.sheet_format.defaultRowHeight or 15
    width_chars = max(int(ws.column_dimensions[col_letter].width or float(target_width)), 1)
    for row, value in values_by_row.items():
        visual_lines = sum(max(1, int(math.ceil(display_width(line) / width_chars))) for line in value.splitlines() or [""])
        want_height = float(base_height) * visual_lines
        if want_height > (ws.row_dimensions[row].height or 0):
            ws.row_dimensions[row].height = want_height
