from __future__ import annotations

from importlib.resources import as_file
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Literal, cast

import streamlit as st
from openpyxl import load_workbook  # type: ignore[import-untyped]

from fnl_builder.config import InputPaths, PipelineConfig
from fnl_builder.pipeline import run
from fnl_builder.render.excel import default_template_ref


def _norm(text: object) -> str:
    return str(text or "").strip().lower().replace(" ", "")


def _find_labeled_value(path: Path, label: str) -> str:
    ws = load_workbook(path, data_only=True).active
    want = _norm(label)
    for r in range(1, 50):
        for c in range(1, 15):
            if _norm(ws.cell(r, c).value) == want:
                return str(ws.cell(r, c + 1).value or "")
    return ""


def _read_guest_rows(path: Path) -> list[dict[str, str]]:
    ws = load_workbook(path, data_only=True).active
    header_row = 0
    for r in range(1, 80):
        vals = [_norm(ws.cell(r, c).value) for c in range(1, 15)]
        if "問い合わせno" in vals:
            header_row = r
            break
    if header_row == 0:
        return []
    rows: list[dict[str, str]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        room_type = str(ws.cell(r, 2).value or "")
        number = str(ws.cell(r, 3).value or "")
        inquiry = str(ws.cell(r, 5).value or "")
        family = str(ws.cell(r, 6).value or "")
        given = str(ws.cell(r, 8).value or "")
        remarks = str(ws.cell(r, 10).value or "")
        if not any([room_type, number, inquiry, family, given, remarks]):
            continue
        rows.append(
            {
                "room_type": room_type,
                "number": number,
                "inquiry": inquiry,
                "family_name": family,
                "given_name": given,
                "remarks": remarks,
            }
        )
    return rows


st.set_page_config(page_title="fnl-builder", layout="wide")
st.title("fnl-builder Web UI")

_LLM_OPTIONS: list[str] = ["none", "openai", "mock"]

with st.sidebar:
    rl = st.file_uploader("RoomingList PDF", type=["pdf"])
    pl = st.file_uploader("PassengerList PDF", type=["pdf"])
    ml = st.file_uploader("MessageList PDF/CSV", type=["pdf", "csv"])
    llm_provider_str: str = str(st.selectbox("LLM Provider", options=_LLM_OPTIONS, index=0) or "none")
    run_clicked = st.button("実行", type="primary")

if run_clicked:
    missing: list[str] = []
    if rl is None:
        missing.append("RoomingList PDF")
    if pl is None:
        missing.append("PassengerList PDF")
    if ml is None:
        missing.append("MessageList PDF/CSV")
    if missing:
        st.error(f"以下のファイルをアップロードしてください: {', '.join(missing)}")
        st.stop()
    assert rl is not None and pl is not None and ml is not None
    try:
        with TemporaryDirectory() as td:
            tmp = Path(td)
            rl_path = tmp / "rooming.pdf"
            rl_path.write_bytes(rl.getvalue())
            pl_path = tmp / "passenger.pdf"
            pl_path.write_bytes(pl.getvalue())
            ml_ext = ml.name.split(".")[-1].lower()
            ml_path = tmp / f"messagelist.{ml_ext}"
            ml_path.write_bytes(ml.getvalue())
            with as_file(default_template_ref()) as template_path:
                out_path = tmp / "final_list.xlsx"
                cfg = PipelineConfig(
                    llm_provider=cast(Literal["none", "openai", "mock"], llm_provider_str),
                    input_paths=InputPaths(
                        rooming=rl_path,
                        passenger=pl_path,
                        messagelist=ml_path,
                        template=template_path,
                        output=out_path,
                        audit=tmp / "final_list_audit.json",
                    ),
                )
                result = run(cfg)
            tour_ref = _find_labeled_value(result.output_path, "Tour REF")
            tour_name = _find_labeled_value(result.output_path, "Tour NAME")
            total_pax = result.audit.counts.total_guests
            c1, c2, c3 = st.columns(3)
            c1.metric("tour_ref", tour_ref)
            c2.metric("tour_name", tour_name)
            c3.metric("total_pax", total_pax)
            st.dataframe(_read_guest_rows(result.output_path), use_container_width=True)
            for issue in result.audit.issues:
                msg = f"[{issue.code}] {issue.message}"
                if issue.level == "error":
                    st.error(msg)
                elif issue.level == "warning":
                    st.warning(msg)
                else:
                    st.info(msg)
            xlsx_bytes = result.output_path.read_bytes()
            st.download_button(
                "final_list.xlsx をダウンロード",
                data=xlsx_bytes,
                file_name="final_list.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as exc:
        st.error(f"実行中にエラーが発生しました: {exc}")
